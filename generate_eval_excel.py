"""
Generates Prompt_Eval_Framework_v2.xlsx
Three tabs: Eval Data | Instructions | Dashboard
"""
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, CellIsRule, FormulaRule
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel

wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════════════════════════════════════
# COLOUR PALETTE
# ═══════════════════════════════════════════════════════════════════════════════
C = {
    'bg_dark':    '0F172A',
    'bg_mid':     '1E293B',
    'bg_card':    '1A2744',
    'purple':     '7C3AED',
    'purple_lt':  'A78BFA',
    'green':      '16A34A',
    'green_lt':   '4ADE80',
    'amber':      'F59E0B',
    'amber_lt':   'FCD34D',
    'red':        'DC2626',
    'slate':      '94A3B8',
    'white':      'FFFFFF',
    'text_dim':   '64748B',
    'header_a':   '1D4ED8',   # section A (blue) — Identification
    'header_b':   '0F766E',   # section B (teal) — Structure
    'header_c':   '6D28D9',   # section C (violet) — Track & Style
    'header_d':   '9D174D',   # section D (rose) — Illustration
    'header_e':   '78350F',   # section E (brown) — Intensity
    'header_f':   '374151',   # section F (charcoal) — Prompts
    'header_g':   '4C1D95',   # section G (deep purple) — Gemini Eval
    'header_h':   '064E3B',   # section H (dark green) — Scoring
    'row_even':   'F8FAFC',
    'row_odd':    'FFFFFF',
}

def fill(hex_color):
    return PatternFill(fill_type='solid', fgColor=hex_color)

def font(bold=False, color='111827', size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name='Calibri')

def border_thin():
    s = Side(border_style='thin', color='CBD5E1')
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium():
    s = Side(border_style='medium', color='475569')
    return Border(left=s, right=s, top=s, bottom=s)

def align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 1 — EVAL DATA
# ═══════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = 'Eval Data'
ws.sheet_view.showGridLines = False
ws.freeze_panes = 'F3'

# ── Column definitions ──────────────────────────────────────────────────────
# Each entry: (header, width, section_color, notes, col_type)
# col_type: 'auto' | 'manual' | 'formula' | 'dropdown'
COLUMNS = [
    # Identification
    ('Unique Code',          26, 'header_a', 'Auto from wizard export (DDMM-ARCH-TRACK-LAYOUT-…)', 'auto'),
    ('Timestamp',            16, 'header_a', 'Auto from wizard',                                   'auto'),
    ('Evaluator',            14, 'header_a', 'Your name',                                          'manual'),
    ('Platform Tested',      16, 'header_a', 'Select from dropdown',                               'dropdown'),
    # Structure
    ('Archetype',            10, 'header_b', 'A1 / A2 / B1',                                      'auto'),
    ('Frame / Overlay Type', 18, 'header_b', 'A1: F1–F5 | B1: O1–O5 | A2: —',                    'auto'),
    ('Layout',               10, 'header_b', 'L1–L6',                                              'auto'),
    ('Overlay Strength (B1)',18, 'header_b', 'Light / Heavy — B1 only',                            'auto'),
    # Track & Style
    ('Track',                16, 'header_c', 'A–E',                                                'auto'),
    ('Colour Palette',       16, 'header_c', 'A1/A2 only',                                        'auto'),
    ('Background Finish',    18, 'header_c', 'A1/A2 only',                                        'auto'),
    ('Render Style',         18, 'header_c', 'A1/A2 only',                                        'auto'),
    # Illustration / Scene
    ('Illustration / Scene Main', 22, 'header_d', 'Main illustration (A1/A2) or main scene element (B1)', 'auto'),
    ('Illus Size (A1)',      14, 'header_d', 'Small / Medium / Large — A1 only',                  'auto'),
    ('Frame Decoration',     16, 'header_d', 'Plain / Simple / Detailed — A1/A2 only',            'auto'),
    ('Motif',                16, 'header_d', 'A1/A2 only',                                        'auto'),
    ('Accents / Supporting', 20, 'header_d', 'A2: accent count | B1: supporting scene elements',  'auto'),
    ('Mood (B1)',             14, 'header_d', 'B1 only — Golden Warm, Misty Soft, etc.',           'auto'),
    # Intensity
    ('Intensity',            16, 'header_e', 'Preset name (Basic / Medium / Complex / etc.)',      'auto'),
    # Prompts
    ('Prompt — Firefly (w/ Text)', 40, 'header_f', 'Full prompt — ≤950 chars',                   'auto'),
    ('Prompt — Firefly (No Text)', 40, 'header_f', 'Full prompt — ≤950 chars',                   'auto'),
    ('Prompt — Midjourney',  40, 'header_f', 'Full prompt with --ar flags',                       'auto'),
    ('Prompt — Leonardo',    40, 'header_f', 'Full prompt — ≤1400 chars',                         'auto'),
    ('Prompt — Gemini',      40, 'header_f', 'Full prompt — no char limit',                       'auto'),
    # Gemini Eval
    ('AI Eval Prompt',        50, 'header_g', 'Auto from wizard — paste + image into Gemini / Claude / ChatGPT to get a score', 'auto'),
    ('AI Model Response',    50, 'header_g', '← Paste the AI model\'s response here (score + issues)',                          'manual'),
    # Scoring
    ('① Overall Score (1–5)', 16, 'header_h', '1=Unusable 2=Poor 3=OK 4=Good 5=Ship it',         'manual'),
    ('② Main Element (Y/P/N)', 16, 'header_h', 'Y=Yes P=Partial N=Missing',                       'dropdown'),
    ('② Supporting (Y/P/N)', 15, 'header_h', 'Y=Yes P=Partial N=Missing/N/A',                    'dropdown'),
    ('② Track Style (Y/N)',   14, 'header_h', 'Does it match the intended track?',                 'dropdown'),
    ('② Mood/Light (Y/N)',    14, 'header_h', 'Lighting + colour matches selection?',              'dropdown'),
    ('② Text Zone (Y/N)',     14, 'header_h', 'Clear uncluttered text area?',                     'dropdown'),
    ('③ Colour Harmony (1–5)', 16, 'header_h', '1=Clashing 3=OK 5=Beautiful',                    'manual'),
    ('③ WhatsApp-Ready (Y/N)', 16, 'header_h', 'Would a real Indian user share this?',            'dropdown'),
    ('④ Failure Tag',         20, 'header_h', 'Primary failure reason (select from list)',         'dropdown'),
    ('⑤ Notes',              30, 'header_h', 'Free text — what went wrong or right',              'manual'),
]

SECTION_LABELS = {
    'header_a': 'IDENTIFICATION',
    'header_b': 'TEMPLATE STRUCTURE',
    'header_c': 'TRACK & STYLE',
    'header_d': 'ILLUSTRATION / SCENE',
    'header_e': 'INTENSITY',
    'header_f': 'GENERATED PROMPTS',
    'header_g': 'GEMINI EVALUATION',
    'header_h': 'QA SCORING — PART B',
}

# ── Row 1: Section label bands ──────────────────────────────────────────────
section_ranges = {}
current_section = None
section_start = None

for ci, (hdr, w, sec, note, ctype) in enumerate(COLUMNS, start=1):
    if sec != current_section:
        if current_section is not None:
            section_ranges[current_section] = (section_start, ci - 1)
        current_section = sec
        section_start = ci
section_ranges[current_section] = (section_start, len(COLUMNS))

# Write section bands on row 1 and set widths
for sec, (col_s, col_e) in section_ranges.items():
    label = SECTION_LABELS.get(sec, '')
    cell = ws.cell(row=1, column=col_s, value=label)
    cell.fill = fill(C[sec])
    cell.font = font(bold=True, color=C['white'], size=9)
    cell.alignment = align(h='center')
    # Merge section label across its columns
    if col_e > col_s:
        ws.merge_cells(start_row=1, start_column=col_s, end_row=1, end_column=col_e)

# ── Row 2: Column headers ────────────────────────────────────────────────────
for ci, (hdr, w, sec, note, ctype) in enumerate(COLUMNS, start=1):
    cell = ws.cell(row=2, column=ci, value=hdr)
    cell.fill = fill('1E293B')
    cell.font = font(bold=True, color=C['white'], size=9)
    cell.alignment = align(h='center', wrap=True)
    cell.border = border_thin()
    ws.column_dimensions[get_column_letter(ci)].width = w

ws.row_dimensions[1].height = 18
ws.row_dimensions[2].height = 36

# ── Row 3: Notes/helper row ──────────────────────────────────────────────────
for ci, (hdr, w, sec, note, ctype) in enumerate(COLUMNS, start=1):
    cell = ws.cell(row=3, column=ci, value=note)
    cell.fill = fill('2D3748')
    cell.font = font(italic=True, color=C['slate'], size=8)
    cell.alignment = align(wrap=True)
    cell.border = border_thin()
ws.row_dimensions[3].height = 20

# ── Pre-format data rows 4–203 ───────────────────────────────────────────────
for row in range(4, 204):
    row_fill = C['row_even'] if row % 2 == 0 else C['row_odd']
    for ci in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=row, column=ci)
        cell.fill = fill(row_fill)
        cell.font = font(size=9)
        cell.alignment = align(wrap=True)
        cell.border = border_thin()

# ── Highlight Overall Score column (col AK = 27) with conditional formatting ─
score_col = 27  # ① Overall Score
score_col_letter = get_column_letter(score_col)
ws.conditional_formatting.add(
    f'{score_col_letter}4:{score_col_letter}203',
    ColorScaleRule(
        start_type='num', start_value=1, start_color='DC2626',
        mid_type='num',   mid_value=3,   mid_color='F59E0B',
        end_type='num',   end_value=5,   end_color='16A34A',
    )
)
# Same for Colour Harmony (col 33)
ch_col = get_column_letter(33)
ws.conditional_formatting.add(
    f'{ch_col}4:{ch_col}203',
    ColorScaleRule(
        start_type='num', start_value=1, start_color='DC2626',
        mid_type='num',   mid_value=3,   mid_color='F59E0B',
        end_type='num',   end_value=5,   end_color='16A34A',
    )
)

# ── Data validations ─────────────────────────────────────────────────────────
def add_dv(ws, formula, col_idx, start_row=4, end_row=203):
    dv = DataValidation(type='list', formula1=formula, allow_blank=True, showErrorMessage=False)
    col_letter = get_column_letter(col_idx)
    dv.add(f'{col_letter}{start_row}:{col_letter}{end_row}')
    ws.add_data_validation(dv)
    return dv

# Platform (col 4)
add_dv(ws, '"Midjourney,Leonardo.AI,Adobe Firefly (Text),Adobe Firefly (No Text),Google Gemini"', 4)
# Y/P/N cols
for ci, (hdr, w, sec, note, ctype) in enumerate(COLUMNS, start=1):
    if ctype == 'dropdown':
        if 'Y/P/N' in hdr:
            add_dv(ws, '"Y,P,N"', ci)
        elif 'Y/N' in hdr:
            add_dv(ws, '"Y,N"', ci)
        elif 'Failure' in hdr:
            add_dv(ws, '"None,Missing main element,Supporting element missing,Wrong track style,Colour mismatch,Text zone blocked,Frame structure wrong,Illustration size wrong,Lighting/mood mismatch,Bad render style,Motif placement error,Photography not realistic,AI text visible,Artifacts/distortion,Too busy,Too sparse,Other"', ci)
        elif 'WhatsApp' in hdr:
            add_dv(ws, '"Y,N"', ci)

# ── Freeze panes after col E (structural), fix row 1+2+3 ────────────────────
ws.freeze_panes = 'E4'

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 2 — INSTRUCTIONS
# ═══════════════════════════════════════════════════════════════════════════════
wi = wb.create_sheet('Instructions')
wi.sheet_view.showGridLines = False
wi.column_dimensions['A'].width = 4
wi.column_dimensions['B'].width = 22
wi.column_dimensions['C'].width = 70

def ins_head(ws, row, text, bg='1D4ED8'):
    ws.merge_cells(f'B{row}:C{row}')
    c = ws.cell(row=row, column=2, value=text)
    c.fill = fill(bg)
    c.font = Font(bold=True, color='FFFFFF', size=11, name='Calibri')
    c.alignment = align(h='left')
    ws.row_dimensions[row].height = 22

def ins_row(ws, row, label, value, label_bg='1E293B', val_bg='F8FAFC'):
    c1 = ws.cell(row=row, column=2, value=label)
    c1.fill = fill(label_bg)
    c1.font = Font(bold=True, color='E2E8F0', size=9, name='Calibri')
    c1.alignment = align(h='left', wrap=True)
    c2 = ws.cell(row=row, column=3, value=value)
    c2.fill = fill(val_bg)
    c2.font = Font(size=9, color='111827', name='Calibri')
    c2.alignment = align(h='left', wrap=True)
    ws.row_dimensions[row].height = 18

r = 2
wi.merge_cells(f'B{r}:C{r}')
title = wi.cell(row=r, column=2, value='Prompt Evaluation Framework — Instructions v2')
title.fill = fill('0F172A')
title.font = Font(bold=True, color='A78BFA', size=14, name='Calibri')
title.alignment = align(h='left')
wi.row_dimensions[r].height = 28

r = 4
ins_head(wi, r, '📋 HOW TO USE — STEP BY STEP', '1D4ED8')
steps = [
    ('Step 1 — Generate',  'Open the HTML wizard. Walk through all steps and generate prompts.'),
    ('Step 2 — Export Row','Click 📋 Export Eval Row. Go to Eval Data tab → click cell A4 (or next empty row) → Paste (Cmd+V). All auto-fill columns populate.'),
    ('Step 3 — Platform',  'Fill col D (Platform Tested) from dropdown: which AI platform you ran the prompt on.'),
    ('Step 4 — Run Prompt','Paste the platform-specific prompt from the Excel into Midjourney / Leonardo / Firefly / Gemini. Save the generated image.'),
    ('Step 5 — AI Eval',    'Col Y (AI Eval Prompt) is auto-filled when you export the row. Copy it → open Gemini / Claude / ChatGPT → paste the prompt → attach the generated image → send. The prompt already has the Firefly generation prompt embedded. If you tested a different platform, paste that platform\'s prompt (from cols T–X) in place of it.'),
    ('Step 6 — Paste Response', 'Copy the AI model\'s full response. Paste it into col Z (AI Model Response) for that row.'),
    ('Step 7 — Score',     'Fill col AA (Overall Score 1–5) manually. Also fill element check dropdowns (cols AB–AF) and col AH (WhatsApp-Ready).'),
    ('Step 8 — Repeat',    'Same combo tested on 4 platforms = 4 rows (one per platform). Paste the same row 4 times, change col D each time.'),
]
for label, val in steps:
    r += 1
    ins_row(wi, r, label, val)

r += 2
ins_head(wi, r, '🔑 UNIQUE CODE — FORMAT GUIDE', '0F766E')
code_rows = [
    ('Format',            'DDMM-[ARCH+STEP2]-[TRACK]-[LAYOUT]-[PAL]-[ILLUS]-[EXTRAS]-[INTENSITY]'),
    ('A1 example',        '2904-A1F2-B-L3-C2-SCC-Si-Md-Me'),
    ('  → decoded',       '29 Apr · A1 Dual Band frame · Modern Desi · Text Left · Soft Dawn · Steaming Chai Cup · Simple deco · Medium size · Medium intensity'),
    ('A2 example',        '2904-A2-D-L1-C5-LD-1ac-Bl'),
    ('  → decoded',       '29 Apr · A2 · Artsy Playful · Text Top · Sacred Saffron · Lit Diya · 1 accent prop · Balanced intensity'),
    ('B1 example',        '2904-B1O2-E-L1-SCC-M1-Lt-Me'),
    ('  → decoded',       '29 Apr · B1 Gradient-Top overlay · Soft & Warm · Text Top · Steaming Chai Cup · Golden Warm mood · Light overlay · Medium intensity'),
    ('ARCH+STEP2 codes',  'A1F1–F5 | A2 (no step2 ID) | B1O1–O5'),
    ('TRACK codes',       'A=Shiny Maximal  B=Modern Desi  C=Minimalist  D=Artsy Playful  E=Soft & Warm'),
    ('INTENSITY abbr',    'Ba=Basic  Me/Bl=Medium/Balanced  Co/Ex/Mx/Ri=Complex variants  Mn=Minimal'),
    ('FRAME DECO abbr',   'Pl=Plain  Si=Simple  Dt=Detailed'),
    ('ILLUS SIZE abbr',   'Sm=Small  Md=Medium  Lg=Large'),
    ('MOOD abbr (B1)',     'M1=Golden Warm  M2=Misty Soft  M3=Bright & Vivid  M4=Dramatic Rays  M5=Vintage Film  M6=Soft Pastel'),
    ('OVERLAY STR (B1)',  'Lt=Light  Hv=Heavy'),
    ('ILLUS abbr',        'First letter of each word in illustration name, uppercase, max 4 chars'),
]
for label, val in code_rows:
    r += 1
    ins_row(wi, r, label, val)

r += 2
ins_head(wi, r, '📊 SCORING GUIDE — PART B', '064E3B')
score_rows = [
    ('① Overall Score (1–5)', '1=Unusable (major failure)  2=Poor (needs significant rework)  3=Decent (usable with tweaks)  4=Good (minor issues)  5=Ship it (production ready)'),
    ('② Main Element (Y/P/N)', 'Y=Clearly visible and well-rendered  P=Partial (distorted/blurry/partially wrong)  N=Missing entirely'),
    ('② Supporting (Y/P/N)',  'Same scale. Mark N/A as N if no supporting elements were selected.'),
    ('② Track Style (Y/N)',   'Does the overall aesthetic feel match the selected track? (e.g. Modern Desi = muted earthy editorial)'),
    ('② Mood/Light (Y/N)',    'Does the lighting + colour grading match the mood selection? (B1 primary; A1/A2 = colour palette match)'),
    ('② Text Zone (Y/N)',     'Is there a visually clear, uncluttered area suitable for text overlay in the expected position?'),
    ('③ Colour Harmony (1–5)', '1=Colours clash badly  2=Somewhat off  3=Acceptable  4=Good harmony  5=Beautiful, cohesive palette'),
    ('③ WhatsApp-Ready (Y/N)', 'Honest gut check: would a real Indian WhatsApp user forward this as a Good Morning greeting?'),
    ('④ Failure Tag',         'Select the PRIMARY reason for failure (or None). Use Notes for more detail.'),
    ('⑤ Notes',              'Free text. What specifically went wrong? Any prompt tweak ideas? What worked well?'),
]
for label, val in score_rows:
    r += 1
    ins_row(wi, r, label, val)

r += 2
ins_head(wi, r, '💡 TIPS FOR QA EFFICIENCY', '9D174D')
tips = [
    ('Batch testing',     'Test all 5 platforms for one combo before moving to the next. Paste same row 5 times, change col D.'),
    ('Filter by archetype', 'Use col E (Archetype) filter to compare A1 vs A2 vs B1 results separately.'),
    ('Sort by score',     'Sort col AA (Overall Score) ascending to find worst combos fast.'),
    ('Filter by Failure Tag', 'Filter col AI to group systemic issues (e.g. all "Text zone blocked" rows).'),
    ('AI eval tip',       'In Gemini / Claude / ChatGPT: paste the AI Eval Prompt (col Y) as text, then attach the image in the SAME message. If you tested Midjourney/Leonardo/Gemini, swap the embedded Firefly prompt with the relevant prompt from cols U/V/W/X before sending.'),
    ('One row = one test', 'Same wizard selections + different platform = different row. This lets you compare platform performance.'),
]
for label, val in tips:
    r += 1
    ins_row(wi, r, label, val)

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 3 — DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════
wd = wb.create_sheet('Dashboard')
wd.sheet_view.showGridLines = False
wd.column_dimensions['A'].width = 3
wd.column_dimensions['B'].width = 24
wd.column_dimensions['C'].width = 14
wd.column_dimensions['D'].width = 14
wd.column_dimensions['E'].width = 14
wd.column_dimensions['F'].width = 14
wd.column_dimensions['G'].width = 14

def dash_title(ws, row, text, col_s=2, col_e=7, bg='0F172A', fg='A78BFA', size=13):
    ws.merge_cells(start_row=row, start_column=col_s, end_row=row, end_column=col_e)
    c = ws.cell(row=row, column=col_s, value=text)
    c.fill = fill(bg)
    c.font = Font(bold=True, color=fg, size=size, name='Calibri')
    c.alignment = align(h='left')
    ws.row_dimensions[row].height = 24

def dash_sub(ws, row, text, col_s=2, col_e=7, bg='1E293B', fg='CBD5E1'):
    ws.merge_cells(start_row=row, start_column=col_s, end_row=row, end_column=col_e)
    c = ws.cell(row=row, column=col_s, value=text)
    c.fill = fill(bg)
    c.font = Font(bold=True, color=fg, size=10, name='Calibri')
    c.alignment = align(h='left')
    ws.row_dimensions[row].height = 18

def dash_hdr(ws, row, cols_vals, col_start=2, bg='2D3748', fg='94A3B8'):
    for ci, val in enumerate(cols_vals, start=col_start):
        c = ws.cell(row=row, column=ci, value=val)
        c.fill = fill(bg)
        c.font = Font(bold=True, color=fg, size=9, name='Calibri')
        c.alignment = align(h='center')
        ws.row_dimensions[row].height = 16

def dash_row(ws, row, label, formulas, col_start=2, label_bg='1E293B', val_bg='F8FAFC', label_fg='E2E8F0', val_fg='111827'):
    c = ws.cell(row=row, column=col_start, value=label)
    c.fill = fill(label_bg)
    c.font = Font(bold=True, color=label_fg, size=9, name='Calibri')
    c.alignment = align(h='left')
    for ci, fml in enumerate(formulas, start=col_start + 1):
        cell = ws.cell(row=row, column=ci, value=fml)
        cell.fill = fill(val_bg)
        cell.font = Font(size=9, color=val_fg, name='Calibri')
        cell.alignment = align(h='center')
    ws.row_dimensions[row].height = 16

# Header note about formula conventions
# AA in Eval Data = col 27 = Overall Score
# E in Eval Data = Archetype (col 5)
# I in Eval Data = Track (col 9)
# D in Eval Data = Platform (col 4)
# R in Eval Data = Illustration/Scene Main (col 18... wait, let me recount)
# Let me map all columns:
# A=1 UniqueCode, B=2 Timestamp, C=3 Evaluator, D=4 Platform,
# E=5 Archetype, F=6 FrameOverlay, G=7 Layout, H=8 OverlayStrength,
# I=9 Track, J=10 Palette, K=11 BGFinish, L=12 RenderStyle,
# M=13 IllusMain, N=14 IllusSize, O=15 FrameDeco, P=16 Motif,
# Q=17 AccentsSupporting, R=18 Mood,
# S=19 Intensity,
# T=20 Prompt-FF-Text, U=21 Prompt-FF-NoText, V=22 Prompt-MJ, W=23 Prompt-Leo, X=24 Prompt-Gemini,
# Y=25 GeminiEvalPrompt, Z=26 GeminiResponse,
# AA=27 OverallScore, AB=28 MainElement, AC=29 Supporting, AD=30 TrackStyle,
# AE=31 Mood/Light, AF=32 TextZone, AG=33 ColourHarmony,
# AH=34 WhatsAppReady, AI=35 FailureTag, AJ=36 Notes

# Data range: 'Eval Data'!AA4:AA203 for scores, E for archetype, I for track, D for platform
SCORE = "'Eval Data'!AA$4:AA$203"
ARCH  = "'Eval Data'!E$4:E$203"
TRACK = "'Eval Data'!I$4:I$203"
PLAT  = "'Eval Data'!D$4:D$203"
WA    = "'Eval Data'!AH$4:AH$203"
MAIN  = "'Eval Data'!AB$4:AB$203"
COLOR_H = "'Eval Data'!AG$4:AG$203"

r = 2
dash_title(wd, r, '📊  Prompt Eval Dashboard', bg='0F172A', fg='A78BFA', size=14)

r = 4
dash_title(wd, r, 'BY ARCHETYPE', bg='1D4ED8', fg='FFFFFF', size=10)
r += 1
dash_hdr(wd, r, ['Archetype', 'Avg Score', 'Runs', 'WA-Ready %', 'Main Hit %', 'Colour Avg'])
for arch_label, arch_val in [('A1 — Text + Frame', 'A1'), ('A2 — Illustrations', 'A2'), ('B1 — Photo BG', 'B1')]:
    r += 1
    score_f  = f'=IFERROR(AVERAGEIF({ARCH},"{arch_val}",{SCORE}),"—")'
    count_f  = f'=COUNTIF({ARCH},"{arch_val}")'
    wa_f     = f'=IFERROR(COUNTIFS({ARCH},"{arch_val}",{WA},"Y")/COUNTIF({ARCH},"{arch_val}")*100,"—")'
    main_f   = f'=IFERROR(COUNTIFS({ARCH},"{arch_val}",{MAIN},"Y")/COUNTIF({ARCH},"{arch_val}")*100,"—")'
    color_f  = f'=IFERROR(AVERAGEIF({ARCH},"{arch_val}",{COLOR_H}),"—")'
    dash_row(wd, r, arch_label, [score_f, count_f, wa_f, main_f, color_f])

r += 2
dash_title(wd, r, 'BY TRACK', bg='6D28D9', fg='FFFFFF', size=10)
r += 1
dash_hdr(wd, r, ['Track', 'Avg Score', 'Runs', 'WA-Ready %', 'Main Hit %', 'Colour Avg'])
tracks = [('A — Shiny Maximal', 'A*'), ('B — Modern Desi', 'B*'), ('C — Minimalist Classy', 'C*'), ('D — Artsy Playful', 'D*'), ('E — Soft & Warm', 'E*')]
for tlabel, tval in tracks:
    r += 1
    # Track ID is just the letter — use wildcard to cover track names
    # Actually track stored as full name in col I. Let me adjust.
    # The eval data exports track name (not ID). So let's match by name prefix.
    track_names = {
        'A*': 'Shiny Maximal', 'B*': 'Modern Desi',
        'C*': 'Minimalist Classy', 'D*': 'Artsy Playful', 'E*': 'Soft & Warm'
    }
    tname = track_names.get(tval, tval.rstrip('*'))
    score_f  = f'=IFERROR(AVERAGEIF({TRACK},"*{tname}*",{SCORE}),"—")'
    count_f  = f'=COUNTIF({TRACK},"*{tname}*")'
    wa_f     = f'=IFERROR(COUNTIFS({TRACK},"*{tname}*",{WA},"Y")/COUNTIF({TRACK},"*{tname}*")*100,"—")'
    main_f   = f'=IFERROR(COUNTIFS({TRACK},"*{tname}*",{MAIN},"Y")/COUNTIF({TRACK},"*{tname}*")*100,"—")'
    color_f  = f'=IFERROR(AVERAGEIF({TRACK},"*{tname}*",{COLOR_H}),"—")'
    dash_row(wd, r, tlabel, [score_f, count_f, wa_f, main_f, color_f])

r += 2
dash_title(wd, r, 'BY PLATFORM', bg='0F766E', fg='FFFFFF', size=10)
r += 1
dash_hdr(wd, r, ['Platform', 'Avg Score', 'Runs', 'WA-Ready %', 'Main Hit %', 'Colour Avg'])
platforms = [
    ('Midjourney',                'Midjourney'),
    ('Leonardo.AI',               'Leonardo.AI'),
    ('Adobe Firefly (w/ Text)',   'Adobe Firefly (Text)'),
    ('Adobe Firefly (No Text)',   'Adobe Firefly (No Text)'),
    ('Google Gemini',             'Google Gemini'),
]
for plabel, pval in platforms:
    r += 1
    score_f  = f'=IFERROR(AVERAGEIF({PLAT},"*{pval}*",{SCORE}),"—")'
    count_f  = f'=COUNTIF({PLAT},"*{pval}*")'
    wa_f     = f'=IFERROR(COUNTIFS({PLAT},"*{pval}*",{WA},"Y")/COUNTIF({PLAT},"*{pval}*")*100,"—")'
    main_f   = f'=IFERROR(COUNTIFS({PLAT},"*{pval}*",{MAIN},"Y")/COUNTIF({PLAT},"*{pval}*")*100,"—")'
    color_f  = f'=IFERROR(AVERAGEIF({PLAT},"*{pval}*",{COLOR_H}),"—")'
    dash_row(wd, r, plabel, [score_f, count_f, wa_f, main_f, color_f])

r += 2
dash_title(wd, r, 'BY MOOD — B1 ONLY', bg='78350F', fg='FFFFFF', size=10)
r += 1
dash_hdr(wd, r, ['Mood', 'Avg Score', 'Runs', 'WA-Ready %', 'Main Hit %', ''])
MOOD = "'Eval Data'!R$4:R$203"
moods = ['Golden Warm', 'Misty Soft', 'Bright & Vivid', 'Dramatic Rays', 'Vintage Film', 'Soft Pastel']
for mood_name in moods:
    r += 1
    score_f = f'=IFERROR(AVERAGEIF({MOOD},"*{mood_name}*",{SCORE}),"—")'
    count_f = f'=COUNTIF({MOOD},"*{mood_name}*")'
    wa_f    = f'=IFERROR(COUNTIFS({MOOD},"*{mood_name}*",{WA},"Y")/COUNTIF({MOOD},"*{mood_name}*")*100,"—")'
    main_f  = f'=IFERROR(COUNTIFS({MOOD},"*{mood_name}*",{MAIN},"Y")/COUNTIF({MOOD},"*{mood_name}*")*100,"—")'
    dash_row(wd, r, mood_name, [score_f, count_f, wa_f, main_f, ''])

r += 2
dash_title(wd, r, 'TOP FAILURE TAGS (counts)', bg='9D174D', fg='FFFFFF', size=10)
r += 1
dash_hdr(wd, r, ['Failure Tag', 'Count', '', '', '', ''])
FAIL = "'Eval Data'!AI$4:AI$203"
failure_tags = [
    'Missing main element', 'Supporting element missing', 'Wrong track style',
    'Colour mismatch', 'Text zone blocked', 'Frame structure wrong',
    'Illustration size wrong', 'Lighting/mood mismatch', 'Bad render style',
    'AI text visible', 'Artifacts/distortion',
]
for tag in failure_tags:
    r += 1
    count_f = f'=COUNTIF({FAIL},"*{tag}*")'
    dash_row(wd, r, tag, [count_f, '', '', '', ''])

r += 2
dash_sub(wd, r, '💡 Tip: Refresh formulas after pasting new data. Use Cmd+Shift+F9 (Mac) or Ctrl+Alt+F9 (Win).')

# ── Apply border to all written cells ────────────────────────────────────────
for sheet in [wd, wi]:
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                cell.border = border_thin()

# ── Tab colours ──────────────────────────────────────────────────────────────
ws.sheet_properties.tabColor   = '16A34A'   # green
wi.sheet_properties.tabColor   = '1D4ED8'   # blue
wd.sheet_properties.tabColor   = '7C3AED'   # purple

# ── Save ─────────────────────────────────────────────────────────────────────
output_path = '/Users/rohan.kohli/Documents/Github/9. Template Automation V2/Prompt_Eval_Framework_v2.xlsx'
wb.save(output_path)
print(f'Saved: {output_path}')
